import pandas as pd
import re
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import PieChart, Reference

def clean_sprint(sprint_str):
    if pd.isna(sprint_str):
        return ""
    sprints = str(sprint_str).split(',')
    clean_sprints = [s.strip() for s in sprints if s.strip()]
    return ", ".join(clean_sprints)

def auto_adjust_columns(worksheet, df):
    """Auto-adjust column widths for readability."""
    for idx, col in enumerate(df.columns):
        series = df[col]
        max_len = max(
            int(series.fillna("").astype(str).map(len).max()) if not series.empty else 0,
            len(str(series.name))
        ) + 2
        max_len = min(max_len, 50)
        col_letter = get_column_letter(idx + 1)
        worksheet.column_dimensions[col_letter].width = max_len

def style_header(worksheet, num_cols, color='4472C4', header_row=1):
    """Apply styling to the header row."""
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

def style_epic_row(worksheet, row_num, num_cols):
    """Highlight Epic rows in a sheet."""
    epic_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    epic_font = Font(bold=True, size=11)
    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(row=row_num, column=col_idx)
        cell.fill = epic_fill
        cell.font = epic_font

def style_task_row(worksheet, row_num, num_cols):
    """Slightly highlight Task rows."""
    task_fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')
    task_font = Font(bold=True, size=10)
    for col_idx in range(1, num_cols + 1):
        cell = worksheet.cell(row=row_num, column=col_idx)
        cell.fill = task_fill
        cell.font = task_font

def add_issue_type_dropdown(worksheet, issue_type_col_idx, num_rows, issue_types):
    """Add a drop-down on the Issue Type column."""
    formula_str = '"' + ','.join(issue_types) + '"'
    dv = DataValidation(type='list', formula1=formula_str, allow_blank=True)
    dv.error = 'Please select a valid Issue Type'
    dv.errorTitle = 'Invalid Issue Type'
    dv.prompt = 'Select an Issue Type'
    dv.promptTitle = 'Issue Type'
    col_letter = get_column_letter(issue_type_col_idx)
    cell_range = f"{col_letter}2:{col_letter}{num_rows + 1}"
    dv.add(cell_range)
    worksheet.add_data_validation(dv)

def safe_sheet_name(name):
    """Escape invalid characters for Excel sheet names and limit to 31 chars."""
    safe_name = str(name).replace("/", "-").replace("\\\\", "-").replace("?", "").replace("*", "").replace("[", "(").replace("]", ")")
    # Also fallback with regex just to be sure
    safe_name = re.sub(r'[\\/*?\[\]:]', '-', safe_name)
    return safe_name[:31]

def build_epic_hierarchy(df_all):
    """Build a dict: epic_key -> { epic_row, tasks: [ {task_row, subtasks: [...]} ] }"""
    epics = df_all[df_all['Issue Type'] == 'Epic']
    tasks = df_all[df_all['Issue Type'].isin(['Task', 'Test'])]
    subtasks = df_all[df_all['Issue Type'] == 'Subtask']
    
    hierarchy = {}
    for _, epic in epics.iterrows():
        epic_key = epic['Issue key']
        # Find tasks that belong to this epic
        child_tasks = tasks[tasks['Parent Key'] == epic_key]
        
        task_list = []
        for _, task in child_tasks.iterrows():
            task_key = task['Issue key']
            # Find subtasks that belong to this task
            child_subtasks = subtasks[subtasks['Parent Key'] == task_key]
            task_list.append({
                'task': task,
                'subtasks': child_subtasks
            })
        
        hierarchy[epic_key] = {
            'epic': epic,
            'tasks': task_list
        }
    
    return hierarchy

def convert_jira_csv_to_excel(input_csv, output_excel):
    print(f"Reading data from {input_csv}...")
    df = pd.read_csv(input_csv)
    
    columns_to_keep = [
        'Issue key', 'Summary', 'Issue Type', 'Status', 'Priority',
        'Assignee', 'Reporter', 'Created', 'Due date',
        'Custom field (Story point estimate)', 'Sprint',
        'Parent key', 'Parent summary'
    ]
    
    available_columns = [col for col in columns_to_keep if col in df.columns]
    df_filtered = df[available_columns].copy()
    
    rename_mapping = {
        'Custom field (Story point estimate)': 'Story Points',
        'Parent key': 'Parent Key',
        'Parent summary': 'Parent Summary'
    }
    df_filtered.rename(columns=rename_mapping, inplace=True)
    
    if 'Sprint' in df_filtered.columns:
        df_filtered['Sprint'] = df_filtered['Sprint'].apply(clean_sprint)
    if 'Created' in df_filtered.columns:
        df_filtered['Created'] = pd.to_datetime(df_filtered['Created'], format='mixed', errors='coerce').dt.strftime('%Y-%m-%d')
    if 'Due date' in df_filtered.columns:
        df_filtered['Due date'] = pd.to_datetime(df_filtered['Due date'], format='mixed', errors='coerce').dt.strftime('%Y-%m-%d')
    if 'Sprint' in df_filtered.columns:
        df_filtered.sort_values(by=['Sprint', 'Status'], inplace=True)

    issue_types = df_filtered['Issue Type'].dropna().unique().tolist() if 'Issue Type' in df_filtered.columns else []
    print(f"Found issue types: {issue_types}")

    # Build epic hierarchy
    hierarchy = build_epic_hierarchy(df_filtered)
    print(f"Found {len(hierarchy)} Epics with hierarchy")

    print(f"Writing data to {output_excel}...")
    
    try:
        with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
            # ========== Sheet 1: Summary ==========
            summary_data = []
            overall_total = 0
            overall_completed = 0
            
            for epic_key, data in hierarchy.items():
                epic_name = str(data['epic']['Summary'])
                
                # Count tasks and subtasks for this epic
                total_tasks = 0
                completed_tasks = 0
                
                for task_data in data['tasks']:
                    total_tasks += 1
                    if task_data['task']['Status'] in ('Done', 'Closed', 'Resolved'):
                        completed_tasks += 1
                        
                    for _, subtask in task_data['subtasks'].iterrows():
                        total_tasks += 1
                        if subtask['Status'] in ('Done', 'Closed', 'Resolved'):
                            completed_tasks += 1
                
                overall_total += total_tasks
                overall_completed += completed_tasks
                
                # Calculate progress
                percent_complete = f"{(completed_tasks / total_tasks * 100):.1f}%" if total_tasks > 0 else "0.0%"
                
                summary_data.append({
                    'Epic': epic_name,
                    'Total Tasks (incl. subtasks)': total_tasks,
                    'Completed Tasks': completed_tasks,
                    'Percent Complete': percent_complete,
                    'Epic Status': data['epic']['Status']
                })
            
            # --- Overall Progress Summary ---
            overall_remaining = overall_total - overall_completed
            overall_percent = f"{(overall_completed / overall_total * 100):.1f}%" if overall_total > 0 else "0.0%"
            
            df_overall = pd.DataFrame([
                {"Metric": "Completed Tasks", "Count": overall_completed},
                {"Metric": "Remaining Tasks", "Count": overall_remaining}
            ])
            
            df_overall.to_excel(writer, index=False, sheet_name='Summary', startrow=0, startcol=0)
            
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, index=False, sheet_name='Summary', startrow=5, startcol=0)
            
            ws_summary = writer.sheets['Summary']
            auto_adjust_columns(ws_summary, df_summary) # adjust column widths based on df_summary
            style_header(ws_summary, len(df_overall.columns), color='548235', header_row=1)
            style_header(ws_summary, len(df_summary.columns), color='4472C4', header_row=6)
            
            ws_summary.auto_filter.ref = f"A6:{get_column_letter(len(df_summary.columns))}{len(df_summary) + 6}"
            ws_summary.freeze_panes = 'A7'
            
            # --- Pie Chart ---
            pie = PieChart()
            labels = Reference(ws_summary, min_col=1, min_row=2, max_row=3)
            data = Reference(ws_summary, min_col=2, min_row=1, max_row=3)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = f"Overall Progress: {overall_percent}"
            ws_summary.add_chart(pie, "D1")
            
            # ========== Per-Epic Sheets (Epic -> Tasks -> Subtasks) ==========
            for epic_key, data in hierarchy.items():
                epic_row = data['epic']
                
                # 🔴 IMPORTANT FIX: Safe Sheet Names
                epic_name = str(epic_row['Summary'])
                sheet_name = safe_sheet_name(epic_name)
                
                # Build a flat dataframe: Epic row, then Task rows, then Subtask rows under each Task
                rows = []
                rows.append(epic_row)
                
                for task_data in data['tasks']:
                    rows.append(task_data['task'])
                    for _, subtask in task_data['subtasks'].iterrows():
                        rows.append(subtask)
                
                if len(rows) == 0:
                    continue
                
                df_epic = pd.DataFrame(rows, columns=df_filtered.columns)
                df_epic.to_excel(writer, index=False, sheet_name=sheet_name)
                
                ws_epic = writer.sheets[sheet_name]
                auto_adjust_columns(ws_epic, df_epic)
                style_header(ws_epic, len(df_epic.columns), color='2E75B6')
                ws_epic.auto_filter.ref = ws_epic.dimensions
                ws_epic.freeze_panes = 'A2'
                
                # Color-code rows by type and add outline grouping
                num_cols = len(df_epic.columns)
                issue_type_list = df_epic['Issue Type'].tolist() if 'Issue Type' in df_epic.columns else []
                
                # We start grouping from level 1 (Tasks) and level 2 (Subtasks/Tests)
                for row_idx, it in enumerate(issue_type_list):
                    excel_row = row_idx + 2  # +2 because row 1 is header, data starts at row 2
                    
                    if it == 'Epic':
                        style_epic_row(ws_epic, excel_row, num_cols)
                        # Epics are outline level 0 (not collapsed)
                    elif it in ('Task', 'Test'):
                        style_task_row(ws_epic, excel_row, num_cols)
                        # Tasks are outline level 1 (collapse under Epics)
                        ws_epic.row_dimensions[excel_row].outlineLevel = 1
                    elif it == 'Subtask':
                        # Subtasks are outline level 2 (collapse under Tasks)
                        ws_epic.row_dimensions[excel_row].outlineLevel = 2
                
                # Configure sheet outline settings
                ws_epic.sheet_properties.outlinePr.summaryBelow = False
                
                total_children = len(rows) - 1
                print(f"  -> Sheet '{sheet_name}': 1 Epic + {total_children} children")

            print(f"\\nSuccessfully converted to {output_excel}")
        print(f"Sheets: {len(hierarchy)} Epic sheets")
    except PermissionError:
        print(f"\\n❌ ERROR: Permission denied for file '{output_excel}'")
        print("Please CLOSE the Excel file if it is open in another program, then run this file again.")
        print("Windows prevents modifying a file while it is currently open in Excel.")

if __name__ == "__main__":
    input_file = 'Jira.csv'
    output_file = 'Jira_Project_Management.xlsx'
    convert_jira_csv_to_excel(input_file, output_file)
